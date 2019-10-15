# coding:utf-8
import os
import xlrd
from openpyxl import load_workbook

from src.config import global_param


class ExcelUtils:
    """对excel处理操作类"""
    p = os.path.abspath('.')

    def __init__(self, book_name, sheet_name):
        book_path = os.path.join(self.p, "data\\"+global_param.language+"\\"+book_name)
        self.sheet = sheet_name
        self.wb = load_workbook(book_path)
        self.table = xlrd.open_workbook(book_path).sheet_by_name(sheet_name)

    def load_excel_by_path(self, file_path):
        return load_workbook(file_path)

    def get_cell_value_by_sheet_name_and_index(self,wb,sheet_name,col_index, row_index):
        ws = wb[sheet_name]
        return ws[col_index+row_index].value

    def get_cell_value(self, col_index, row_index):
        ws = self.wb[self.sheet]
        return ws[col_index+row_index].value

    def dict_data(self):
        nrows = self.table.nrows
        nclos = self.table.ncols

        if nrows <= 1:
            print("总行数小于1")
        else:
            r = []
            j = 1
            for i in list(range(nrows - 1)):
                s = {}
                # 从第二行取对应values值
                # s['rowNum'] = i+2
                values = self.table.row_values(j)
                for x in list(range(nclos)):
                    s[self.table.row_values(0)[x]] = values[x]
                r.append(s)
                j += 1
            return r



