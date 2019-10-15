# encoding=utf-8
import traceback
from datetime import datetime

import time
from openpyxl import load_workbook
from selenium import webdriver

import xlsxwriter  # 往Excel表格写数据 安装cmd输入： pip install XlsxWriter，前置条件已安装pip

import xlrd, os  # xlrd模块用于读EXCEL文档数据，安装xlrd在cmd中输入pip install xlrd，前置条件已安装pip

# 新建Excel表格
from xlrd import xldate_as_tuple

from config import Var
from config.Var import project_path
from unit.DirAndFile import createDir


def write_content():
    # 新建Excel表格
    testreport = project_path + '/test_excel/'
    dirPath = createDir(testreport, '')
    os.chdir(dirPath)
    xl = xlsxwriter.Workbook(testreport + '/login_info.xlsx')  # 新建Excel表格，注意需要在当前目录下新建testData文件夹
    table = xl.add_worksheet('result1')  # 新建一个名为result的页签
    # 设置格式：行高、列宽、底色、字号等格式
    blue = xl.add_format({'font_name': 'Arial', 'font_size': 10, 'color': 'blue'})
    red = xl.add_format({'font_name': 'Arial', 'font_size': 10, 'color': 'red'})
    # 设置第一行的行高为30
    table.set_row(0, 15)
    # 设置第1列列宽为50
    table.set_column(0, 3, 30)
    # 第2列到第3列的列宽为25
    table.set_column(1, 2, 25)
    # 写数据
    table.write_string(0, 0, '名称', blue)
    table.write_string(0, 1, '用户名', red)  # 往第一行第一列写入字符串数据
    table.write_string(0, 2, '密码', red)
    table.write_string(0, 3, '预期结果', red)

    table.write_string(1, 0, 'user_null', blue)  # 往第二行第一列写入数字,本地搭建的bugfree地址
    table.write_string(1, 1, '', red)
    table.write_string(1, 2, '123', red)
    table.write_string(1, 3, '用户名/密码不能为空', red)

    # 关闭Excel表格
    xl.close()


# 读取数据
# def read_content():
#     # 一下脚本实现bugfree登录，测试数据从Excel读取
#     testreport = project_path + '/test_excel/'
#     data = xlrd.open_workbook(testreport + '/login_info.xlsx')
#     tableGet = data.sheet_by_name('result1')
#     testData1 = tableGet.row_values(1)  # 读取第二行数据
#     data = testData1[0], testData1[1], testData1[2], testData1[3]
#     testData2 = tableGet.row_values(2)  # 读取第二行数据
#
#     data1 = testData2[0], testData2[1], testData2[2], testData2[3]
#     testData3 = tableGet.row_values(3)  # 读取第二行数据
#
#     data2 = testData3[0], testData3[1], testData3[2], testData3[3]
#
#     return [data, data1, data2]


# def read_excel( filename, sheetname):
#     rbook = xlrd.open_workbook(filename)
#     sheet = rbook.sheet_by_name(sheetname)
#     rows = sheet.nrows
#     cols = sheet.ncols
#     all_content = []
#     for i in range(rows):
#         row_content = []
#         for j in range(cols):
#             ctype = sheet.cell(i, j).ctype  # 表格的数据类型
#             cell = sheet.cell_value(i, j)
#             if ctype == 2 and cell % 1 == 0:  # 如果是整形
#                 cell = int(cell)
#             elif ctype == 3:
#                 # 转成datetime对象
#                 date = datetime(*xldate_as_tuple(cell, 0))
#                 cell = date.strftime('%Y/%d/%m %H:%M:%S')
#             elif ctype == 4:
#                 cell = True if cell == 1 else False
#             row_content.append(cell)
#         all_content.append(tuple(row_content))
#         # print('[' + ','.join("'" + str(element) + "'" for element in row_content) + ']')
#     return all_content[1:]
class ExcelUtils:
    """对excel处理操作类"""

    def __init__(self, book_name, sheet_name):
        testreport = project_path + '/test_excel/'
        book_path = testreport + book_name
        self.sheet = sheet_name
        self.wb = load_workbook(book_path)
        self.table = xlrd.open_workbook(book_path).sheet_by_name(sheet_name)

    def load_excel_by_path(self, file_path):
        return load_workbook(file_path)

    def get_cell_value_by_sheet_name_and_index(self, wb, sheet_name, col_index, row_index):
        ws = wb[sheet_name]
        return ws[col_index + row_index].value

    def get_cell_value(self, col_index, row_index):
        ws = self.wb[self.sheet]
        return ws[col_index + row_index].value

    def dict_data(self):
        nrows = self.table.nrows
        nclos = self.table.ncols

        if nrows <= 1:
            print("总行数小于1")
        else:
            r = []
            j = 1
            for i in list(range(nrows - 1)):
                s = {}
                # 从第二行取对应values值
                # s['rowNum'] = i+2
                values = self.table.row_values(j)
                for x in list(range(nclos)):
                    s[self.table.row_values(0)[x]] = values[x]
                r.append(s)
                j += 1
            return r


def read_excel(filename, sheetname):
    rbook = xlrd.open_workbook(filename)
    sheet = rbook.sheet_by_name(sheetname)
    # 行
    rows = sheet.nrows
    # 列
    cols = sheet.ncols
    all_content = []
    for i in range(rows):
        row_content = []
        for j in range(cols):
            ctype = sheet.cell(i, j).ctype  # 表格的数据类型
            cell = sheet.cell_value(i, j)
            if ctype == 2 and cell % 1 == 0:  # 如果是整形
                cell = int(cell)
            elif ctype == 3:
                # 转成datetime对象
                date = datetime(*xldate_as_tuple(cell, 0))
                cell = date.strftime('%Y/%d/%m %H:%M:%S')
            elif ctype == 4:
                cell = True if cell == 1 else False
            row_content.append(cell)
        all_content.append(tuple(row_content))
        # print('[' + ','.join("'" + str(element) + "'" for element in row_content) + ']')
    return all_content[1:]


if __name__ == '__main__':
    filename = 'login_info.xlsx'
    sheetname = 'Sheet1'
    a = ExcelUtils(filename, sheetname)
    print(a.dict_data())

# 启动脚本
# if __name__ == '__main__':
#     # 数据的写入
#     # write_content()
#     # # 数据的读取
#     tup_result = read_content()
#     print(tup_result)
#     # 启动脚本去执行
#     # driver = webdriver.Firefox()
#     # bugFreeLogin(driver, tup_result[0], tup_result[1], tup_result[2])


